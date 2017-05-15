import cv2
import numpy as np
from openpyxl import Workbook

# for ms excel e.g. 'AZ' to 'BA'
def NextColumn(string):
    # ['A', 'B', 'C', .. etc]
    alphaB = [chr(asciiNum) for asciiNum in range(65, 91)]

    # extract alphabet from string
    onlyAlpha = ""
    for char in string:
        if char.isalpha():
            onlyAlpha += char

    # if last character in 'onlyAlpha' is  n o t  'Z'
    if alphaB.index(onlyAlpha[-1]) + 1 <= alphaB.index('Z'):
        # go to next letter e.g. 'A' to 'B'
        onlyAlpha = onlyAlpha[:-1] + alphaB[alphaB.index(onlyAlpha[-1]) + 1]

    # if 'Z'
    else:

        # change last character from 'Z' to 'A'
        onlyAlpha = onlyAlpha[:-1] + 'A'

        # assume all onlyAlpha characters from first character to previous last character [:-1]
        # are 'Z'(s) >> if so, they will be change to 'A'(s) and a new 'A' will be added to the end
        isZs = True

        # loop through from from first character to previous last character [:-1]
        for pos in range(len(onlyAlpha) - 2, -1, -1):
            # if 'Z'
            if alphaB.index(onlyAlpha[pos]) + 1 > alphaB.index('Z') and isZs == True:
                onlyAlpha = onlyAlpha[:pos] + 'A' + onlyAlpha[pos + 1:]

            # if not 'Z'
            else:
                if isZs == True:
                    # go to next letter e.g. 'A' to 'B' >> only once can happened >> if it happened, no more flips
                    onlyAlpha = onlyAlpha[:pos] + alphaB[alphaB.index(onlyAlpha[pos]) + 1] + onlyAlpha[pos + 1:]
                isZs = False

        # only in case all characters in onlyAlpha is/are 'Z'(s)
        if isZs == True:
            onlyAlpha = onlyAlpha + 'A'

    return onlyAlpha

# for debug (let pixels be more readable)
# record pixels in excel file >> each pixel column in text photo should be recorded in excel column
def photoTextPixelToExcel(text, by_width_or_height):
    wb = Workbook()
    ws = wb.active

    # prev: will be used to compare the    w i d t h   of previous iteration with current one
    # + here store the   w i d t h  of first element in array as first element in iteration
    prev = int(text[0][by_width_or_height])

    # store the first element in excel
    ws['A1'] = str(text[0][0]) + ' ' + str(+ text[0][1]) + ' ' + str(text[0][2])

    currentExcelLetter = 'A'
    currentExcelNum = 2

    # loop from the element that is from 2nd element
    for x in range(1, text.size):
        current = int(text[x][by_width_or_height])
        if (current != prev):
            # get the next excel comlumn
            currentExcelLetter = NextColumn(currentExcelLetter)
            currentExcelNum = 1
        ws[currentExcelLetter + str(currentExcelNum)] = str(text[x][0]) + ' ' + str(+ text[x][1]) + ' ' + str(text[x][2])
        currentExcelNum += 1
        prev = current

    wb.save("empty_book.xlsx")

# extract   w h o l e   text pixels from photo : sorted by rows
def extractWholeTextFromPhoto(img, img_height, img_width):
    # *** img should be already converted to black and white

    dtype = np.dtype([('height', int, 1), ('width', int, 1), ('color', int, 1)])
    text = np.array([(0, 0, 0)], dtype=dtype)

    #img[columns, width]
    # searching pixels
    # 1 . top to bottom     >>      range(img_height)
    # 2 . left to right     >>      range(img_width)

    for x in range(img_width):
        for y in range(img_height):

            # check if it is is character (recognize non whited spaces)
            if img[y, x] != 255:

                # store characters
                aText = np.array([(y, x, img[y, x])], dtype=dtype)
                text = np.concatenate((text, aText), axis=0)

    text = np.delete(text, 0, 0)
    return text

# extract text borders' pixels from photo : sorted by rows
def extractTextBorderFromPhotoByWidth(img, img_height, img_width):

    dtype = np.dtype([('height', int, 1), ('width', int, 1), ('color', int, 1)])
    text = np.array([(0, 0, 0)], dtype=dtype)


    #img[columns, width]
    # searching pixels
    # 1 . top to bottom     >>      range(img_height)
    # 2 . left to right     >>      range(img_width)

    for x in range(img_width):
        # True: while the text pixels start on a column only store the first pixel
        # and set isText to True and never store another pixel until isText become False
        isText = True

        # the first pixel on column
        prev = img[0, x]

        # check if first pixel on column is character (recognize non whited spaces)
        if prev != 255:
            # store characters
            aText = np.array([(0, x, img[0, x])], dtype=dtype)
            text = np.concatenate((text,aText), axis=0)
            isText = False

        for y in range(1, img_height):
            current = img[y, x]

            # check if it is is character (recognize non whited spaces)
            if current != prev and current != 255 and isText == True:
                # store characters
                aText = np.array([(y, x, img[y, x])], dtype=dtype)
                text = np.concatenate((text, aText), axis=0)
                isText = False
            else:
                # the bottom text border that is followed by white pixel
                if current == 255 and prev != 255:
                    aText = np.array([(y-1, x, prev)], dtype=dtype)
                    text = np.concatenate((text, aText), axis=0)
                    isText = True


            prev = img[y, x]

    text = np.delete(text, 0, 0)
    return text


# extract text borders' pixels from photo : sorted by rows
def extractTextBorderFromPhotoByHeight(img, img_height, img_width):

    dtype = np.dtype([('height', int, 1), ('width', int, 1), ('color', int, 1)])
    text = np.array([(0, 0, 0)], dtype=dtype)


    #img[columns, width]
    # searching pixels
    # 1 . left to right     >>      range(img_width)
    # 2 . top to bottom     >>      range(img_height)

    for y in range(img_height):
        # True: while the text pixels start on a column only store the first pixel
        # and set isText to True and never store another pixel until isText become False
        isText = True

        # the first pixel on column
        prev = img[y, 0]

        # check if first pixel on column is character (recognize non whited spaces)
        if prev != 255:
            # store characters
            aText = np.array([(y, 0, img[y, 0])], dtype=dtype)
            text = np.concatenate((text,aText), axis=0)
            isText = False

        for x in range(1, img_width):
            current = img[y, x]

            # check if it is is character (recognize non whited spaces)
            if current != prev and current != 255 and isText == True:
                # store characters
                aText = np.array([(y, x, img[y, x])], dtype=dtype)
                text = np.concatenate((text, aText), axis=0)
                isText = False
            else:
                # the bottom text border that is followed by white pixel
                if current == 255 and prev != 255:
                    aText = np.array([(y, x-1, prev)], dtype=dtype)
                    text = np.concatenate((text, aText), axis=0)
                    isText = True


            prev = img[y, x]

    text = np.delete(text, 0, 0)
    return text


# extract an inner text pixels from photo
def extractInnerLineByWidth(img, img_text):

    for border in range(0, img_text.size, 2):
        top_border = img_text[border][0]
        bottom_border = img_text[border + 1][0]


        border_difference = bottom_border - (top_border - 1) # - 1 because ::: e.g. top_border is 83 and bottom_border is 85
                                                # 85 - 83 is 2 !! while the total difference is 83 (1) + 84 (1) + 85(1)
                                                # = 3
        half = border_difference / 2
        #print(top_border, bottom_border, border_difference, half, end="")

        if border_difference % 2 == 0:
            half += 1
        else:
            half += 0.5

        half = int(half)

        img[(top_border - 1) + half][img_text[border][1]] = 180

        #print("",half)
    return img

# extract an inner text pixels from photo
def extractInnerLineByHeight(img, img_text):

    for border in range(0, img_text.size, 2):
        top_border = img_text[border][1]
        bottom_border = img_text[border + 1][1]


        border_difference = bottom_border - (top_border - 1) # - 1 because ::: e.g. top_border is 83 and bottom_border is 85
                                                # 85 - 83 is 2 !! while the total difference is 83 (1) + 84 (1) + 85(1)
                                                # = 3
        half = border_difference / 2
        #print(top_border, bottom_border, border_difference, half, end="")

        if border_difference % 2 == 0:
            half += 1
        else:
            half += 0.5

        half = int(half)

        img[img_text[border][0]][(top_border - 1) + half] = 80

        #print("",half)
    return img


def main():
    np.set_printoptions(threshold=np.nan)

    img = cv2.imread('test.png',cv2.IMREAD_GRAYSCALE)
    height, width = img.shape

    img_text = extractTextBorderFromPhotoByHeight(img, height, width)
    img = extractInnerLineByHeight(img, img_text)

    img_text = extractTextBorderFromPhotoByWidth(img, height, width)
    img = extractInnerLineByWidth(img, img_text)

    # order by_height = 0
    # order by_width = 1
    #photoTextPixelToExcel(img_text, 0)

    print(img_text)
    cv2.imwrite('result.png', img)
    cv2.imshow('image',img)
    cv2.waitKey(0)
    cv2.destroyAllWindows()

main()